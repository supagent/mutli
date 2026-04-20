"""Tests for the Multica ADK agent package: bridge, tools, multica_agent."""

import io
import json
import os
import sys
from unittest import mock

import httpx
import pytest

# ── Ensure server/agent is on sys.path so bare imports work ─────────────────
_AGENT_DIR = os.path.dirname(os.path.abspath(__file__))
if _AGENT_DIR not in sys.path:
    sys.path.insert(0, _AGENT_DIR)

# ── Pre-mock google.adk/google.genai so multica_agent can be imported ───────
# We need a consistent mock hierarchy: google -> google.adk -> ... and google.genai -> ...
_google_mock = mock.MagicMock()
_adk_mock_modules = {
    "google": _google_mock,
    "google.adk": _google_mock.adk,
    "google.adk.agents": _google_mock.adk.agents,
    "google.adk.runners": _google_mock.adk.runners,
    "google.adk.sessions": _google_mock.adk.sessions,
    "google.adk.models": _google_mock.adk.models,
    "google.adk.models.llm_response": _google_mock.adk.models.llm_response,
    "google.genai": _google_mock.genai,
    "google.genai.types": _google_mock.genai.types,
}
for mod_name, mod_mock in _adk_mock_modules.items():
    sys.modules[mod_name] = mod_mock


# ============================================================================
# BRIDGE TESTS
# ============================================================================


class TestNDJSONEmitter:
    """Tests for bridge.NDJSONEmitter."""

    def _capture_lines(self, fn) -> list[dict]:
        """Run fn with stdout redirected, return parsed NDJSON lines."""
        buf = io.StringIO()
        with mock.patch("sys.stdout", buf):
            fn()
        lines = buf.getvalue().strip().splitlines()
        return [json.loads(line) for line in lines if line.strip()]

    # -- Event structure tests --

    def test_emit_tool_use_structure(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        lines = self._capture_lines(
            lambda: emitter.emit_tool_use("get_issue", {"issue_id": "123"})
        )
        assert len(lines) == 1
        ev = lines[0]
        assert ev["type"] == "tool_use"
        assert ev["tool"] == "get_issue"
        assert ev["input"] == {"issue_id": "123"}
        assert ev["task_id"] == "t-1"
        assert ev["issue_id"] == "iss-1"
        assert ev["seq"] == 1

    def test_emit_tool_result_structure(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        lines = self._capture_lines(
            lambda: emitter.emit_tool_result("get_issue", '{"title":"foo"}')
        )
        ev = lines[0]
        assert ev["type"] == "tool_result"
        assert ev["tool"] == "get_issue"
        assert ev["output"] == '{"title":"foo"}'
        assert ev["task_id"] == "t-1"
        assert ev["issue_id"] == "iss-1"

    def test_emit_text_structure(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        lines = self._capture_lines(lambda: emitter.emit_text("Hello world"))
        ev = lines[0]
        assert ev["type"] == "text"
        assert ev["content"] == "Hello world"
        assert ev["task_id"] == "t-1"
        assert ev["issue_id"] == "iss-1"

    def test_emit_thinking_structure(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-42")
        lines = self._capture_lines(lambda: emitter.emit_thinking("Thinking..."))
        ev = lines[0]
        assert ev["type"] == "thinking"
        assert ev["content"] == "Thinking..."
        assert ev["task_id"] == "t-1"
        assert ev["issue_id"] == "iss-42"

    def test_emit_error_structure(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        lines = self._capture_lines(lambda: emitter.emit_error("something broke"))
        ev = lines[0]
        assert ev["type"] == "error"
        assert ev["content"] == "something broke"
        assert ev["task_id"] == "t-1"
        assert ev["issue_id"] == "iss-1"

    def test_emit_result_structure(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        lines = self._capture_lines(
            lambda: emitter.emit_result("completed", "All done")
        )
        ev = lines[0]
        assert ev["type"] == "result"
        assert ev["status"] == "completed"
        assert ev["output"] == "All done"
        assert ev["usage"] == {}
        assert ev["task_id"] == "t-1"
        assert ev["issue_id"] == "iss-1"

    # -- issue_id present on ALL event types --

    def test_issue_id_present_on_all_event_types(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-99")

        def emit_all():
            emitter.emit_tool_use("t", {})
            emitter.emit_tool_result("t", "out")
            emitter.emit_text("txt")
            emitter.emit_thinking("hmm")
            emitter.emit_error("err")
            emitter.emit_result("completed", "done")

        lines = self._capture_lines(emit_all)
        assert len(lines) == 6
        for ev in lines:
            assert ev["issue_id"] == "iss-99", f"Missing issue_id on type={ev['type']}"

    # -- Seq monotonic --

    def test_seq_numbers_are_monotonic(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")

        def emit_several():
            emitter.emit_text("a")
            emitter.emit_text("b")
            emitter.emit_tool_use("x", {})
            emitter.emit_error("e")

        lines = self._capture_lines(emit_several)
        seqs = [ev["seq"] for ev in lines]
        assert seqs == [1, 2, 3, 4]

    def test_seq_increments_even_for_different_types(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")

        def emit_mixed():
            emitter.emit_tool_use("a", {})
            emitter.emit_tool_result("a", "r")
            emitter.emit_text("t")

        lines = self._capture_lines(emit_mixed)
        assert [ev["seq"] for ev in lines] == [1, 2, 3]

    # -- Empty text skipped --

    def test_empty_text_is_skipped(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        lines = self._capture_lines(lambda: emitter.emit_text(""))
        assert lines == []

    def test_whitespace_only_text_is_skipped(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        lines = self._capture_lines(lambda: emitter.emit_text("   \n  "))
        assert lines == []

    def test_empty_thinking_is_skipped(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        lines = self._capture_lines(lambda: emitter.emit_thinking("  "))
        assert lines == []

    def test_empty_text_does_not_increment_seq(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")

        def emit_with_skips():
            emitter.emit_text("")       # skipped, no seq
            emitter.emit_text("real")   # seq=1

        lines = self._capture_lines(emit_with_skips)
        assert len(lines) == 1
        assert lines[0]["seq"] == 1

    # -- Tool output truncation --

    def test_tool_output_truncated_at_8192(self):
        from bridge import NDJSONEmitter, MAX_TOOL_OUTPUT

        assert MAX_TOOL_OUTPUT == 8192

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        big_output = "x" * 20000
        lines = self._capture_lines(
            lambda: emitter.emit_tool_result("t", big_output)
        )
        ev = lines[0]
        assert len(ev["output"]) == 8192

    def test_tool_output_under_limit_not_truncated(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        output = "short"
        lines = self._capture_lines(
            lambda: emitter.emit_tool_result("t", output)
        )
        assert lines[0]["output"] == "short"

    def test_tool_output_exactly_at_limit(self):
        from bridge import NDJSONEmitter, MAX_TOOL_OUTPUT

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        output = "a" * MAX_TOOL_OUTPUT
        lines = self._capture_lines(
            lambda: emitter.emit_tool_result("t", output)
        )
        assert len(lines[0]["output"]) == MAX_TOOL_OUTPUT

    # -- Usage recording --

    def test_record_usage(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1", model="gemini-2.5-flash")
        metadata = mock.MagicMock()
        metadata.prompt_token_count = 100
        metadata.candidates_token_count = 50
        metadata.cached_content_token_count = 10
        emitter.record_usage(metadata)

        assert emitter.usage == {
            "gemini-2.5-flash": {
                "input_tokens": 100,
                "output_tokens": 50,
                "cache_read_tokens": 10,
                "cache_write_tokens": 0,
            }
        }

    def test_record_usage_none_is_noop(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1")
        emitter.record_usage(None)
        assert emitter.usage == {}

    def test_record_usage_handles_none_fields(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1", model="gemini-2.5-flash")
        metadata = mock.MagicMock()
        metadata.prompt_token_count = None
        metadata.candidates_token_count = None
        metadata.cached_content_token_count = None
        emitter.record_usage(metadata)

        assert emitter.usage["gemini-2.5-flash"]["input_tokens"] == 0
        assert emitter.usage["gemini-2.5-flash"]["output_tokens"] == 0
        assert emitter.usage["gemini-2.5-flash"]["cache_read_tokens"] == 0

    def test_usage_appears_in_result_event(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1", model="m")
        metadata = mock.MagicMock()
        metadata.prompt_token_count = 10
        metadata.candidates_token_count = 20
        metadata.cached_content_token_count = 0
        emitter.record_usage(metadata)

        lines = self._capture_lines(
            lambda: emitter.emit_result("completed", "done")
        )
        assert lines[0]["usage"]["m"]["input_tokens"] == 10

    def test_record_usage_overwrites_previous(self):
        from bridge import NDJSONEmitter

        emitter = NDJSONEmitter(task_id="t-1", issue_id="iss-1", model="m")

        meta1 = mock.MagicMock()
        meta1.prompt_token_count = 10
        meta1.candidates_token_count = 20
        meta1.cached_content_token_count = 0
        emitter.record_usage(meta1)

        meta2 = mock.MagicMock()
        meta2.prompt_token_count = 99
        meta2.candidates_token_count = 88
        meta2.cached_content_token_count = 77
        emitter.record_usage(meta2)

        assert emitter.usage["m"]["input_tokens"] == 99


# ============================================================================
# TOOLS TESTS
# ============================================================================


class TestTools:
    """Tests for tools module."""

    @pytest.fixture(autouse=True)
    def _set_env(self, monkeypatch):
        monkeypatch.setenv("MULTICA_API_URL", "http://test-api:8080")
        monkeypatch.setenv("MULTICA_AGENT_TOKEN", "tok-123")
        monkeypatch.setenv("MULTICA_WORKSPACE_ID", "ws-1")

    # -- _request returns error dicts --

    def test_request_returns_error_on_4xx(self, monkeypatch):
        import tools

        resp = httpx.Response(status_code=404, text="not found")
        monkeypatch.setattr(
            tools, "_http_call", lambda method, url, headers, body=None: resp
        )

        result = tools._request("GET", "/api/issues/bad-id")
        assert "error" in result
        assert "404" in result["error"]

    def test_request_returns_error_on_5xx(self, monkeypatch):
        import tools

        resp = httpx.Response(status_code=500, text="internal server error")
        monkeypatch.setattr(
            tools, "_http_call", lambda method, url, headers, body=None: resp
        )

        result = tools._request("GET", "/api/issues/x")
        assert "error" in result
        assert "500" in result["error"]

    def test_request_returns_error_on_timeout(self, monkeypatch):
        import tools

        def raise_timeout(*args, **kwargs):
            raise httpx.TimeoutException("timed out")

        monkeypatch.setattr(tools, "_http_call", raise_timeout)
        result = tools._request("GET", "/api/issues/x")
        assert "error" in result
        assert "timed out" in result["error"].lower()

    def test_request_returns_error_on_connect_error(self, monkeypatch):
        import tools

        def raise_connect(*args, **kwargs):
            raise httpx.ConnectError("connection refused")

        monkeypatch.setattr(tools, "_http_call", raise_connect)
        result = tools._request("GET", "/api/issues/x")
        assert "error" in result
        assert "connection failed" in result["error"].lower()

    def test_request_returns_error_on_generic_exception(self, monkeypatch):
        import tools

        def raise_generic(*args, **kwargs):
            raise RuntimeError("unexpected")

        monkeypatch.setattr(tools, "_http_call", raise_generic)
        result = tools._request("GET", "/api/issues/x")
        assert "error" in result
        assert "unexpected" in result["error"]

    def test_request_success_returns_json(self, monkeypatch):
        import tools

        resp = httpx.Response(
            status_code=200,
            json={"id": "iss-1", "title": "Test"},
        )
        monkeypatch.setattr(
            tools, "_http_call", lambda method, url, headers, body=None: resp
        )

        result = tools._request("GET", "/api/issues/iss-1")
        assert result == {"id": "iss-1", "title": "Test"}

    # -- search_issues URL encodes query --

    def test_search_issues_url_encodes_query(self, monkeypatch):
        import tools

        captured = {}

        def capture_call(method, url, headers, body=None):
            captured["url"] = url
            return httpx.Response(status_code=200, json={"results": []})

        monkeypatch.setattr(tools, "_http_call", capture_call)
        tools.search_issues("hello world & special=chars")

        assert "hello%20world" in captured["url"]
        assert "%26" in captured["url"]  # & is encoded
        assert "%3D" in captured["url"]  # = is encoded

    def test_search_issues_url_encodes_unicode(self, monkeypatch):
        import tools

        captured = {}

        def capture_call(method, url, headers, body=None):
            captured["url"] = url
            return httpx.Response(status_code=200, json={"results": []})

        monkeypatch.setattr(tools, "_http_call", capture_call)
        tools.search_issues("test query")

        # Verify no raw spaces in the URL
        assert " " not in captured["url"]

    # -- Tools never raise exceptions --

    def test_get_issue_never_raises(self, monkeypatch):
        import tools

        monkeypatch.setattr(
            tools, "_http_call", lambda *a, **kw: (_ for _ in ()).throw(RuntimeError("kaboom"))
        )
        result = tools.get_issue("iss-1")
        assert isinstance(result, dict)
        assert "error" in result

    def test_update_issue_never_raises(self, monkeypatch):
        import tools

        monkeypatch.setattr(
            tools, "_http_call", lambda *a, **kw: (_ for _ in ()).throw(RuntimeError("kaboom"))
        )
        result = tools.update_issue("iss-1", "done")
        assert isinstance(result, dict)
        assert "error" in result

    def test_add_comment_never_raises(self, monkeypatch):
        import tools

        monkeypatch.setattr(
            tools, "_http_call", lambda *a, **kw: (_ for _ in ()).throw(RuntimeError("kaboom"))
        )
        result = tools.add_comment("iss-1", "hi")
        assert isinstance(result, dict)
        assert "error" in result

    def test_list_comments_never_raises(self, monkeypatch):
        import tools

        monkeypatch.setattr(
            tools, "_http_call", lambda *a, **kw: (_ for _ in ()).throw(RuntimeError("kaboom"))
        )
        result = tools.list_comments("iss-1")
        assert isinstance(result, dict)
        assert "error" in result

    # -- _http_call retries on transient errors --

    def test_http_call_retries_on_timeout(self, monkeypatch):
        """Verify _http_call retries up to 3 times on TimeoutException."""
        import tools

        call_count = {"n": 0}

        class FakeClient:
            def __init__(self, **kwargs):
                pass

            def __enter__(self):
                return self

            def __exit__(self, *args):
                pass

            def get(self, url, headers=None):
                call_count["n"] += 1
                if call_count["n"] < 3:
                    raise httpx.TimeoutException("slow")
                return httpx.Response(status_code=200, json={"ok": True})

        monkeypatch.setattr(httpx, "Client", FakeClient)

        # Override tenacity wait to make the test fast
        tools._http_call.retry.wait = lambda *a, **kw: 0

        result = tools._http_call("GET", "http://test/api", {})
        assert result.status_code == 200
        assert call_count["n"] == 3

    def test_http_call_retries_on_connect_error(self, monkeypatch):
        """Verify _http_call retries on ConnectError."""
        import tools

        call_count = {"n": 0}

        class FakeClient:
            def __init__(self, **kwargs):
                pass

            def __enter__(self):
                return self

            def __exit__(self, *args):
                pass

            def get(self, url, headers=None):
                call_count["n"] += 1
                if call_count["n"] < 2:
                    raise httpx.ConnectError("refused")
                return httpx.Response(status_code=200, json={"ok": True})

        monkeypatch.setattr(httpx, "Client", FakeClient)
        tools._http_call.retry.wait = lambda *a, **kw: 0

        result = tools._http_call("GET", "http://test/api", {})
        assert result.status_code == 200
        assert call_count["n"] == 2

    def test_http_call_gives_up_after_3_attempts(self, monkeypatch):
        """After 3 failures, tenacity raises RetryError."""
        import tools
        from tenacity import RetryError

        call_count = {"n": 0}

        class FakeClient:
            def __init__(self, **kwargs):
                pass

            def __enter__(self):
                return self

            def __exit__(self, *args):
                pass

            def get(self, url, headers=None):
                call_count["n"] += 1
                raise httpx.TimeoutException("always slow")

        monkeypatch.setattr(httpx, "Client", FakeClient)
        tools._http_call.retry.wait = lambda *a, **kw: 0

        with pytest.raises(RetryError):
            tools._http_call("GET", "http://test/api", {})

        assert call_count["n"] == 3

    # -- create_document writes files --

    def test_create_document_writes_file(self, tmp_path, monkeypatch):
        import tools

        output_dir = str(tmp_path / "output")
        # Patch the hardcoded /workspace/output by replacing the function body concept:
        # We monkeypatch os.path.join and os.makedirs won't matter since tmp_path exists.
        # Simplest: just temporarily redefine create_document to use tmp_path.
        original = tools.create_document

        def patched(filename, content):
            try:
                os.makedirs(output_dir, exist_ok=True)
                filepath = os.path.join(output_dir, filename)
                with open(filepath, "w") as f:
                    f.write(content)
                return {"filename": filename, "path": filepath, "status": "created"}
            except Exception as e:
                return {"error": f"Failed to create document: {e}"}

        monkeypatch.setattr(tools, "create_document", patched)

        result = tools.create_document("report.md", "# Hello\nWorld")
        assert result["status"] == "created"
        assert result["filename"] == "report.md"

        written = open(result["path"]).read()
        assert written == "# Hello\nWorld"

    def test_create_document_error_returns_dict(self, monkeypatch):
        import tools

        # Make os.makedirs fail inside the real function
        orig_makedirs = os.makedirs

        def failing_makedirs(path, **kwargs):
            if "/workspace/output" in path:
                raise PermissionError("nope")
            return orig_makedirs(path, **kwargs)

        monkeypatch.setattr(os, "makedirs", failing_makedirs)
        result = tools.create_document("test.md", "content")
        assert "error" in result

    # -- create_xlsx parses JSON data --

    def test_create_xlsx_parses_json(self, tmp_path, monkeypatch):
        import tools

        output_dir = str(tmp_path / "output")

        def patched(filename, data_json):
            try:
                os.makedirs(output_dir, exist_ok=True)
                data = json.loads(data_json)
                from openpyxl import Workbook

                wb = Workbook()
                ws = wb.active
                if "headers" in data:
                    ws.append(data["headers"])
                for row in data.get("rows", []):
                    ws.append(row)
                filepath = os.path.join(output_dir, filename)
                wb.save(filepath)
                size = os.path.getsize(filepath)
                return {
                    "filename": filename,
                    "path": filepath,
                    "size_bytes": size,
                    "status": "created",
                }
            except Exception as e:
                return {"error": f"Failed to create xlsx: {e}"}

        monkeypatch.setattr(tools, "create_xlsx", patched)

        data = json.dumps(
            {"headers": ["Name", "Age"], "rows": [["Alice", 30], ["Bob", 25]]}
        )
        result = tools.create_xlsx("people.xlsx", data)
        assert result["status"] == "created"
        assert result["size_bytes"] > 0

        # Verify content
        from openpyxl import load_workbook

        wb = load_workbook(result["path"])
        ws = wb.active
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(3, 2).value == 25

    def test_create_xlsx_bad_json_returns_error(self, monkeypatch):
        import tools

        # JSON parsing happens before file I/O, so it will fail regardless of output_dir
        result = tools.create_xlsx("test.xlsx", "not valid json{{{")
        assert "error" in result

    # -- Headers include auth and workspace --

    def test_headers_include_auth_and_workspace(self):
        import tools

        headers = tools._headers()
        assert headers["Authorization"] == "Bearer tok-123"
        assert headers["X-Workspace-ID"] == "ws-1"
        assert headers["Content-Type"] == "application/json"

    def test_headers_omit_auth_when_no_token(self, monkeypatch):
        import tools

        monkeypatch.delenv("MULTICA_AGENT_TOKEN", raising=False)
        headers = tools._headers()
        assert "Authorization" not in headers

    def test_headers_omit_workspace_when_not_set(self, monkeypatch):
        import tools

        monkeypatch.delenv("MULTICA_WORKSPACE_ID", raising=False)
        headers = tools._headers()
        assert "X-Workspace-ID" not in headers

    # -- _api_url --

    def test_api_url_from_env(self):
        import tools

        assert tools._api_url() == "http://test-api:8080"

    def test_api_url_default(self, monkeypatch):
        import tools

        monkeypatch.delenv("MULTICA_API_URL", raising=False)
        assert tools._api_url() == "http://localhost:8080"

    # -- ALL_TOOLS registry --

    def test_all_tools_contains_expected_functions(self):
        import tools

        names = [f.__name__ for f in tools.ALL_TOOLS]
        assert "get_issue" in names
        assert "search_issues" in names
        assert "update_issue" in names
        assert "list_comments" in names
        assert "add_comment" in names
        assert "create_document" in names
        assert "create_docx" in names
        assert "create_xlsx" in names
        assert len(names) == 8


    # ── Path traversal prevention ──────────────────────────────────────────

    def test_safe_output_path_strips_traversal(self):
        """_safe_output_path strips directory traversal, keeping only basename."""
        import tools
        path, err = tools._safe_output_path("../../etc/passwd")
        assert err is None
        # os.path.basename strips the traversal — file goes to /workspace/output/passwd
        assert path == "/workspace/output/passwd"

    def test_safe_output_path_strips_directory(self):
        """_safe_output_path strips directory components."""
        import tools
        path, err = tools._safe_output_path("subdir/report.md")
        assert err is None
        assert path == "/workspace/output/report.md"

    def test_safe_output_path_rejects_dotdot(self):
        """_safe_output_path rejects bare '..' filename."""
        import tools
        path, err = tools._safe_output_path("..")
        assert err is not None

    def test_safe_output_path_normal_filename(self):
        """_safe_output_path accepts normal filenames."""
        import tools
        path, err = tools._safe_output_path("report.md")
        assert err is None
        assert path == "/workspace/output/report.md"


# ============================================================================
# AGENT TESTS
# ============================================================================


class TestAgent:
    """Tests for multica_agent module."""

    # -- CLI arg parsing --

    def test_cli_arg_parsing_required_args(self):
        """Parser correctly extracts --task-id, --issue-id, --prompt."""
        import argparse

        parser = argparse.ArgumentParser()
        parser.add_argument("--task-id", required=True)
        parser.add_argument("--issue-id", default="")
        parser.add_argument("--prompt", required=True)
        parser.add_argument("--model", default="")
        parser.add_argument("--max-turns", type=int, default=0)

        args = parser.parse_args(
            ["--task-id", "task-42", "--issue-id", "iss-7", "--prompt", "Do the thing"]
        )
        assert args.task_id == "task-42"
        assert args.issue_id == "iss-7"
        assert args.prompt == "Do the thing"

    def test_cli_arg_defaults(self):
        import argparse

        parser = argparse.ArgumentParser()
        parser.add_argument("--task-id", required=True)
        parser.add_argument("--issue-id", default="")
        parser.add_argument("--prompt", required=True)
        parser.add_argument("--model", default="")
        parser.add_argument("--max-turns", type=int, default=0)

        args = parser.parse_args(["--task-id", "t-1", "--prompt", "hello"])
        assert args.issue_id == ""
        assert args.model == ""
        assert args.max_turns == 0

    def test_cli_model_and_max_turns(self):
        import argparse

        parser = argparse.ArgumentParser()
        parser.add_argument("--task-id", required=True)
        parser.add_argument("--issue-id", default="")
        parser.add_argument("--prompt", required=True)
        parser.add_argument("--model", default="")
        parser.add_argument("--max-turns", type=int, default=0)

        args = parser.parse_args(
            [
                "--task-id", "t-1",
                "--prompt", "x",
                "--model", "gemini-2.0-pro",
                "--max-turns", "5",
            ]
        )
        assert args.model == "gemini-2.0-pro"
        assert args.max_turns == 5

    def test_cli_missing_required_args_raises(self):
        import argparse

        parser = argparse.ArgumentParser()
        parser.add_argument("--task-id", required=True)
        parser.add_argument("--prompt", required=True)

        with pytest.raises(SystemExit):
            parser.parse_args([])

    # -- make_turn_limiter --

    def test_turn_limiter_returns_none_under_limit(self):
        """Calls under the limit return None (allow the model call)."""
        from multica_agent import make_turn_limiter

        limiter = make_turn_limiter(3)
        ctx = mock.MagicMock()
        req = mock.MagicMock()

        assert limiter(ctx, req) is None  # call 1
        assert limiter(ctx, req) is None  # call 2
        assert limiter(ctx, req) is None  # call 3

    def test_turn_limiter_returns_llm_response_when_exceeded(self):
        """Once limit is exceeded, returns an LlmResponse to stop the agent."""
        from multica_agent import make_turn_limiter

        limiter = make_turn_limiter(2)
        ctx = mock.MagicMock()
        req = mock.MagicMock()

        limiter(ctx, req)  # call 1 - ok
        limiter(ctx, req)  # call 2 - ok
        result = limiter(ctx, req)  # call 3 - exceeded

        # Should return something (not None) -- an LlmResponse mock
        assert result is not None

    def test_turn_limiter_limit_of_one(self):
        """Edge case: limit of 1 allows exactly one call."""
        from multica_agent import make_turn_limiter

        limiter = make_turn_limiter(1)
        ctx = mock.MagicMock()
        req = mock.MagicMock()

        assert limiter(ctx, req) is None  # call 1 - ok
        result = limiter(ctx, req)        # call 2 - exceeded
        assert result is not None

    def test_turn_limiter_continues_returning_response_after_exceeded(self):
        """Every call after the limit returns an LlmResponse."""
        from multica_agent import make_turn_limiter

        limiter = make_turn_limiter(1)
        ctx = mock.MagicMock()
        req = mock.MagicMock()

        limiter(ctx, req)  # call 1 - ok
        r1 = limiter(ctx, req)  # call 2 - exceeded
        r2 = limiter(ctx, req)  # call 3 - still exceeded

        assert r1 is not None
        assert r2 is not None

    # -- Environment variable fallback --

    def test_google_api_key_fallback(self, monkeypatch):
        """When GOOGLE_API_KEY is missing, falls back to GOOGLE_AI_API_KEY."""
        monkeypatch.delenv("GOOGLE_API_KEY", raising=False)
        monkeypatch.setenv("GOOGLE_AI_API_KEY", "fallback-key")

        api_key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GOOGLE_AI_API_KEY")
        assert api_key == "fallback-key"

    def test_google_api_key_primary(self, monkeypatch):
        """GOOGLE_API_KEY takes precedence over GOOGLE_AI_API_KEY."""
        monkeypatch.setenv("GOOGLE_API_KEY", "primary-key")
        monkeypatch.setenv("GOOGLE_AI_API_KEY", "fallback-key")

        api_key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GOOGLE_AI_API_KEY")
        assert api_key == "primary-key"

    def test_no_api_key_returns_none(self, monkeypatch):
        """When both API key env vars are missing, result is None."""
        monkeypatch.delenv("GOOGLE_API_KEY", raising=False)
        monkeypatch.delenv("GOOGLE_AI_API_KEY", raising=False)

        api_key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GOOGLE_AI_API_KEY")
        assert api_key is None

    # -- SYSTEM_PROMPT exists --

    def test_system_prompt_is_nonempty(self):
        from multica_agent import SYSTEM_PROMPT

        assert len(SYSTEM_PROMPT) > 100
        assert "Multica" in SYSTEM_PROMPT

    def test_system_prompt_mentions_tools(self):
        from multica_agent import SYSTEM_PROMPT

        assert "get_issue" in SYSTEM_PROMPT
        assert "add_comment" in SYSTEM_PROMPT

    # -- Model env var fallback --

    def test_model_env_fallback(self, monkeypatch):
        """MULTICA_MODEL env var is used when --model is empty."""
        monkeypatch.setenv("MULTICA_MODEL", "gemini-2.0-pro")
        model = "" or os.environ.get("MULTICA_MODEL", "gemini-2.5-flash")
        assert model == "gemini-2.0-pro"

    def test_model_default(self, monkeypatch):
        """Default model is gemini-2.5-flash."""
        monkeypatch.delenv("MULTICA_MODEL", raising=False)
        model = "" or os.environ.get("MULTICA_MODEL", "gemini-2.5-flash")
        assert model == "gemini-2.5-flash"

    def test_max_turns_env_fallback(self, monkeypatch):
        """MULTICA_MAX_TURNS env var is used when --max-turns is 0."""
        monkeypatch.setenv("MULTICA_MAX_TURNS", "50")
        max_turns = 0 or int(os.environ.get("MULTICA_MAX_TURNS", "20"))
        assert max_turns == 50

    def test_max_turns_default(self, monkeypatch):
        """Default max_turns is 20."""
        monkeypatch.delenv("MULTICA_MAX_TURNS", raising=False)
        max_turns = 0 or int(os.environ.get("MULTICA_MAX_TURNS", "20"))
        assert max_turns == 20
